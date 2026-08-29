from pathlib import Path
from typing import List, Union
import openpyxl

from backend.src.schemas.manifest import RawSheetPayload, StatementType

def load_excel_sheet(file_path: Union[str, Path], sheet_name: str) -> RawSheetPayload:
    """Reads a single worksheet from an Excel workbook and returns a RawSheetPayload."""
    path = Path(file_path)
    wb = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    raw_grid: List[List] = []
    for row in ws.iter_rows(values_only=True):
        raw_grid.append(list(row))

    wb.close()

    row_count = len(raw_grid)
    col_count = max((len(r) for r in raw_grid), default=0)

    return RawSheetPayload(
        source_filename = path.name,
        sheet_name = sheet_name,
        raw_grid = raw_grid,
        row_count = row_count,
        col_count = col_count,
        detected_type = StatementType.UNKNOWN
    )

def load_excel_sources(file_paths: List[Union[str, Path]]) -> List[RawSheetPayload]:
    """Accepts a list of Excel file paths (handles single-sheet files, multi-tab files, or multiple separate workbooks) and extracts all sheets."""
    all_sheets: List[RawSheetPayload] = []

    for file_path in file_paths:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Source file not found: {path}")

        #Load workbook to inspect all available sheet names
        wb = openpyxl.load_workbook(filename=path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        for s_name in sheet_names:
            payload = load_excel_sheet(path, s_name)
            all_sheets.append(payload)

    return all_sheets
